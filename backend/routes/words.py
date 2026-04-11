"""Word routes for managing dictionary words."""

import csv
import io
from io                 import BytesIO, StringIO
from flask              import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from extensions         import db
from models             import Word, Collection, CollectionMember

import requests
import openpyxl


words_bp = Blueprint("words", __name__)


# ── Helper: get definition from Free Dictionary API ───────────────────────────
# Uses https://dictionaryapi.dev — completely free, no key needed
def lookup_word(word: str) -> dict:
    """Looks up a word using the Free Dictionary API and returns its definition, part of speech, and example sentence."""

    url = f"https://api.dictionaryapi.dev/api/v2/entries/en/{word}"
    response = requests.get(url)

    if response.status_code != 200:
        raise ValueError(f"Word '{word}' not found in dictionary")

    data = response.json()

    # the API returns a list of entries — grab the first one
    entry    = data[0]
    meanings = entry.get("meanings", [])

    if not meanings:
        raise ValueError("No meanings found")

    # grab the first meaning and first definition
    first_meaning    = meanings[0]
    part_of_speech   = first_meaning.get("partOfSpeech", "")
    definitions      = first_meaning.get("definitions", [])
    first_definition = definitions[0] if definitions else {}

    return {
        "definition":    first_definition.get("definition", ""),
        "part_of_speech": part_of_speech,
        "example":       first_definition.get("example", ""),
    }


# ── Helper: check collection access ──────────────────────────────────────────
def has_access(user_id, collection_id):
    """Checks if the user has access to the specified collection. 
        Returns (True, collection) if access is granted, 
        or (False, error_message) if denied."""
    
    collection = Collection.query.get(collection_id)
    if not collection:
        return False, "Collection not found"
    # owner always has access
    if collection.owner_id == int(user_id):
        return True, collection
    # shared collection — check membership
    if collection.is_shared:
        member = CollectionMember.query.filter_by(
            collection_id=collection_id,
            user_id=user_id
        ).first()
        if member:
            return True, collection
    return False, "Access denied"


# ── Add a word ────────────────────────────────────────────────────────────────
# POST /api/words/
# Body: { word, notes (optional), collection_id }
@words_bp.route("/", methods=["POST"])
@jwt_required()
def add_word():
    """Adds a new word to a collection. If no collection_id is provided, it defaults to the user's personal collection."""

    user_id = get_jwt_identity()
    data    = request.get_json()

    if not data or "word" not in data:
        return jsonify({"error": "word is required"}), 400

    collection_id = data.get("collection_id")

    # if no collection specified, use the user's personal collection
    if not collection_id:
        personal = Collection.query.filter_by(
            owner_id  = user_id,
            is_shared = False,
            name      = "Personal"
        ).first()
        collection_id = personal.id if personal else None

    # verify access
    ok, result = has_access(user_id, collection_id)
    if not ok:
        return jsonify({"error": result}), 403

    # check for duplicate in this collection
    existing = Word.query.filter(
        Word.collection_id == collection_id,
        db.func.lower(Word.word) == data["word"].strip().lower()
    ).first()
    if existing:
        return jsonify({"error": f'"{data["word"]}" is already in this collection.'}), 409

    # get definition from the dictionary API; proceed without one if it fails
    definition_found = True
    try:
        meaning = lookup_word(data["word"])
    except Exception:
        meaning = {}
        definition_found = False

    # save the word
    word = Word(
        word           = data["word"],
        part_of_speech = meaning.get("part_of_speech"),
        definition     = meaning.get("definition"),
        example        = meaning.get("example"),
        notes          = data.get("notes", ""),
        added_by       = user_id,
        collection_id  = collection_id
    )
    db.session.add(word)
    db.session.commit()

    return jsonify({
        "message":          "Word added",
        "word":             word.to_dict(),
        "definition_found": definition_found,
    }), 201


# ── List words ────────────────────────────────────────────────────────────────
# GET /api/words/?collection_id=<id>
@words_bp.route("/", methods=["GET"])
@jwt_required()
def list_words():
    """Lists all words in a collection. If no collection_id is provided, it defaults to the user's personal collection."""

    user_id = get_jwt_identity()
    collection_id = request.args.get("collection_id")

    # default to personal collection
    if not collection_id:
        personal = Collection.query.filter_by(
            owner_id=user_id, is_shared=False, name="Personal"
        ).first()
        collection_id = personal.id if personal else None

    ok, result = has_access(user_id, collection_id)
    if not ok:
        return jsonify({"error": result}), 403

    words = Word.query.filter_by(collection_id=collection_id)\
                      .order_by(Word.created_at.desc()).all()

    return jsonify({"words": [w.to_dict() for w in words]}), 200


# ── Delete a word ─────────────────────────────────────────────────────────────
# DELETE /api/words/<word_id>
@words_bp.route("/<int:word_id>", methods=["DELETE"])
@jwt_required()
def delete_word(word_id):
    """Deletes a word. Only the user who added the word can delete it."""

    user_id = get_jwt_identity()
    word    = Word.query.get(word_id)

    if not word:
        return jsonify({"error": "Word not found"}), 404

    # only the person who added it can delete it
    if str(word.added_by) != str(user_id):
        return jsonify({"error": "You can only delete words you added"}), 403

    db.session.delete(word)
    db.session.commit()
    return jsonify({"message": "Word deleted"}), 200


# ── Edit word notes (inline) ─────────────────────────────────────────────────
# PATCH /api/words/<word_id>
# Body: { notes }
@words_bp.route("/<int:word_id>", methods=["PATCH"])
@jwt_required()
def update_word(word_id):
    """Updates the notes field of a word. Only the user who added the word can edit it."""
    user_id = get_jwt_identity()
    word    = Word.query.get(word_id)

    if not word:
        return jsonify({"error": "Word not found"}), 404

    if str(word.added_by) != str(user_id):
        return jsonify({"error": "You can only edit words you added"}), 403

    data = request.get_json()
    if data is None:
        return jsonify({"error": "Request body required"}), 400

    word.notes      = data.get("notes", word.notes)
    word.definition = data.get("definition", word.definition)
    db.session.commit()

    return jsonify({"word": word.to_dict()}), 200


# ── Import words from Excel / CSV ─────────────────────────────────────────────
# POST /api/words/import
# Form data: file (.xlsx or .csv), collection_id (optional)
@words_bp.route("/import", methods=["POST"])
@jwt_required()
def import_words():
    """Imports words from an uploaded .xlsx or .csv file into a collection."""
    user_id = get_jwt_identity()

    if "file" not in request.files:
        return jsonify({"error": "file is required"}), 400

    file          = request.files["file"]
    collection_id = request.form.get("collection_id")
    filename      = file.filename.lower()

    # resolve collection
    if not collection_id:
        personal = Collection.query.filter_by(
            owner_id=user_id, is_shared=False, name="Personal"
        ).first()
        collection_id = personal.id if personal else None

    ok, result = has_access(user_id, collection_id)
    if not ok:
        return jsonify({"error": result}), 403

    # normalise header name
    def norm(s):
        return s.strip().lower().replace(" ", "_").replace("-", "_") if s else ""

    def row_to_word(row_dict):
        """Convert a header→value dict to a Word object or None if word cell is empty."""
        w = row_dict.get("word") or row_dict.get("Word") or ""
        if not str(w).strip():
            return None
        word_text      = str(w).strip()
        definition     = str(row_dict.get("definition") or row_dict.get("Definition") or "").strip() or None
        part_of_speech = str(row_dict.get("part_of_speech") or row_dict.get("Part of Speech") or row_dict.get("part of speech") or "").strip() or None
        example        = str(row_dict.get("example") or row_dict.get("Example") or row_dict.get("Example Sentence") or "").strip() or None
        notes          = str(row_dict.get("notes") or row_dict.get("Notes") or "").strip() or None

        # only call the dictionary API if no definition was provided in the file
        if not definition:
            try:
                meaning        = lookup_word(word_text)
                definition     = meaning.get("definition")
                part_of_speech = part_of_speech or meaning.get("part_of_speech")
                example        = example or meaning.get("example")
            except Exception:
                pass  # import word without definition rather than skip entirely

        return Word(
            word           = word_text,
            part_of_speech = part_of_speech,
            definition     = definition,
            example        = example,
            notes          = notes,
            added_by       = user_id,
            collection_id  = collection_id,
        )

    words_to_add = []
    skipped      = 0

    try:
        if filename.endswith(".xlsx"):
            wb   = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return jsonify({"error": "Empty file"}), 400
            headers = [str(c).strip() if c is not None else "" for c in rows[0]]
            for row in rows[1:]:
                row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
                w = row_to_word(row_dict)
                if w:
                    words_to_add.append(w)
                else:
                    skipped += 1

        elif filename.endswith(".csv"):
            content = file.read().decode("utf-8", errors="replace")
            reader  = csv.DictReader(StringIO(content))
            for row_dict in reader:
                w = row_to_word(row_dict)
                if w:
                    words_to_add.append(w)
                else:
                    skipped += 1
        else:
            return jsonify({"error": "Only .xlsx and .csv files are supported"}), 400

    except Exception as e:
        return jsonify({"error": f"Could not parse file: {str(e)}"}), 400

    if not words_to_add:
        return jsonify({"error": "No valid words found in file"}), 400

    db.session.bulk_save_objects(words_to_add)
    db.session.commit()

    return jsonify({"imported": len(words_to_add), "skipped": skipped}), 201


# ── Export to Excel ───────────────────────────────────────────────────────────
# GET /api/words/export?collection_id=<id>
@words_bp.route("/export", methods=["GET"])
@jwt_required()
def export_words():
    """Exports all words in a collection to an Excel file. 
        If no collection_id is provided, it defaults to the user's personal collection."""
    
    user_id = get_jwt_identity()
    collection_id = request.args.get("collection_id")

    if not collection_id:
        personal = Collection.query.filter_by(
            owner_id=user_id, is_shared=False, name="Personal"
        ).first()
        collection_id = personal.id if personal else None

    ok, result = has_access(user_id, collection_id)
    if not ok:
        return jsonify({"error": result}), 403

    words = Word.query.filter_by(collection_id=collection_id)\
                      .order_by(Word.created_at.desc()).all()

    # build Excel
    wb       = openpyxl.Workbook()
    ws       = wb.active
    ws.title = "Vocabulary"

    # header row
    headers = ["Word", "Part of Speech", "Definition", "Example Sentence", "Notes", "Date Added"]
    ws.append(headers)

    # set column widths
    col_widths = [16, 14, 45, 45, 25, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # data rows
    for w in words:
        ws.append([
            w.word,
            w.part_of_speech or "",
            w.definition     or "",
            w.example        or "",
            w.notes          or "",
            w.created_at.strftime("%Y-%m-%d")
        ])

    # stream the file — no need to save to disk
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment = True,
        download_name = "wordvault.xlsx"
    )